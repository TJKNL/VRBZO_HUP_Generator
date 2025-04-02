import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from data_management import get_executable_relative_path
import os


class KRO_Tree:
    def __init__(self, dataframe_aanzien: pd.DataFrame, dataframe_gebruik: pd.DataFrame):
        self.data_aanzien = dataframe_aanzien
        self.data_aanzien['risico_classificatie'] = 'A'  # Add new column with all rows 'A'
        self.original_data = dataframe_aanzien.copy()
        self.data_gebruik = dataframe_gebruik
        self.HUP = pd.DataFrame(columns=dataframe_aanzien.columns)
        self.history = []
        self.filter_functions = {
            ">": lambda df, column, value: df[df[column] > value],
            "<": lambda df, column, value: df[df[column] < value],
            "==": lambda df, column, value: df[df[column] == value],
            "!=": lambda df, column, value: df[df[column] != value],
            ">=": lambda df, column, value: df[df[column] >= value],
            "<=": lambda df, column, value: df[df[column] <= value],
            "in": lambda df, column, value: df[df[column].isin(value)]
        }

    def filter(self, column, operator, value):
        if column == "personen":
            self.filter_personen(operator, value)
        elif operator in self.filter_functions:
            df_filtered = self.filter_functions[operator](self.data_aanzien, column, value)
            rows_removed = len(self.data_aanzien) - len(df_filtered)
            rows_remaining = len(df_filtered)
            print("")
            print(f"Filtering on {column} {operator} {value}")
            print(f"Removed {rows_removed} rows, {rows_remaining} rows remaining.")
            self.data_aanzien = df_filtered
            self.history.append({"action": f"filter {column} {operator} {value}", "rows_removed": rows_removed,
                                 "rows_remaining": rows_remaining})
        else:
            print("Please provide a valid operator.")

    def filter_personen(self, operator, value):
        # List to store the bronsleutel that satisfy the filter condition
        valid_bronsleutel = []

        # Iterate over each row in 'data_aanzien'
        for index, row in self.data_aanzien.iterrows():
            bronsleutel = row['bronsleutel']  # Get the corresponding 'bronsleutel'

            # Check if 'bronsleutel' exists as 'aanzien_id' in 'data_gebruik'
            matched_rows = self.data_gebruik[self.data_gebruik['aanzien_id'] == bronsleutel]
            # If 'bronsleutel' exists in 'data_gebruik', check if any of the 'personen' values satisfy the filter condition
            if not matched_rows.empty:
                if not self.filter_functions[operator](matched_rows, 'personen', float(value)).empty:
                    valid_bronsleutel.append(bronsleutel)
                    # print(f"bronsleutel: {bronsleutel} satisfies the filter condition.")
                else:
                    pass
                    # print(f"bronsleutel: {bronsleutel} does not satisfy the filter condition.")
            else:
                # If there are zero matches for a bronsleutel -> aanzien_id the filter is true
                valid_bronsleutel.append(bronsleutel)
                # print(
                #    f"bronsleutel: {bronsleutel} does not have a match in data_gebruik, so the filter condition is deemed true.")

        # Log the filter operation in history - FIX: use dictionary format like other methods
        original_rows = len(self.data_aanzien)
        self.data_aanzien = self.data_aanzien[self.data_aanzien['bronsleutel'].isin(valid_bronsleutel)]
        filtered_rows = len(self.data_aanzien)
        removed_rows = original_rows - filtered_rows
        print("")
        print(f"Filtering on Personen {operator} {value}")
        
        # Fix: use dictionary format for history
        self.history.append({"action": f"filter Personen {operator} {value}", 
                            "rows_removed": removed_rows, 
                            "rows_remaining": filtered_rows})

        print(f"Removed {removed_rows} rows, {filtered_rows} rows remaining.")

    def filter_or(self, filter1, filter2):
        column1, operator1, value1 = filter1
        column2, operator2, value2 = filter2

        # Check if the operators provided are valid
        if operator1 not in self.filter_functions or operator2 not in self.filter_functions:
            print("Please provide valid operators.")
            return

        # Apply the filters based on the column and operator
        if column1 == "personen":
            df_filtered1 = self.filter_persons(operator1, value1)
        else:
            df_filtered1 = self.filter_functions[operator1](self.data_aanzien, column1, value1)

        if column2 == "personen":
            df_filtered2 = self.filter_persons(operator2, value2)
        else:
            df_filtered2 = self.filter_functions[operator2](self.data_aanzien, column2, value2)

        # Combine the filtered DataFrames and remove duplicates
        df_filtered = pd.concat([df_filtered1, df_filtered2]).drop_duplicates()

        # Calculate the number of rows removed and remaining
        rows_removed = len(self.data_aanzien) - len(df_filtered)
        rows_remaining = len(df_filtered)

        print("")
        print(f"Filtering on {column1} {operator1} {value1} OR {column2} {operator2} {value2}")
        print(f"Removed {rows_removed} rows, {rows_remaining} rows remaining.")

        # Update the data_aanzien attribute with the filtered DataFrame
        self.data_aanzien = df_filtered

        # Append the filter action to the history list
        self.history.append({"action": f"filter {column1} {operator1} {value1} OR {column2} {operator2} {value2}",
                             "rows_removed": rows_removed, "rows_remaining": rows_remaining})

    def apply_filters(self, filters):
        for column, operator, value in filters:
            self.filter(column, operator, value)

    def get_history(self, index=None):
        if index is None:
            return self.history
        elif 0 <= index < len(self.history):
            return self.history[index]
        else:
            print("Please provide a valid index.")
            return None

    def store_results(self):
        """
        Stores the remaining rows of the current DataFrame in the HUP DataFrame.
        """
        self.HUP = pd.concat([self.HUP, self.data_aanzien])
        self.history.append(
            {"action": "store results into HUP", "rows_removed": 0, "rows_remaining": len(self.data_aanzien)})

    def reset(self):
        """
        Resets the DataFrame to its original state, keeping the 'risico_classificatie' column.
        """
        risico_classificatie = self.data_aanzien['risico_classificatie'].copy()
        self.data_aanzien = self.original_data.copy()
        self.data_aanzien['risico_classificatie'] = risico_classificatie

    def save_hup(self):
        """
        Saves the HUP DataFrame to a csv file.
        """
        self.HUP.to_csv(get_executable_relative_path("data", "HUP-all_data.csv"), index=False)
        self.history.append({"action": "save HUP to csv", "rows_removed": 0, "rows_remaining": 0})

    def set_risk(self, risk_class):
        """
        Sets the 'risico_classificatie' for the remaining items in the DataFrame to a given risk classification.
        """
        valid_classes = ["A", "B", "C"]
        if risk_class not in valid_classes:
            print(f"Invalid risk classification. Please provide one of the following: {valid_classes}")
            return

        self.data_aanzien['risico_classificatie'] = risk_class
        self.history.append({"action": f"Set risk classification to {risk_class}",
                             "rows_remaining": len(self.data_aanzien)})
        print(f"Set risk classification to {risk_class}")

    def filter_SBI(self, start_nums):
        # Ensure start_nums is a list
        if isinstance(start_nums, int):
            start_nums = [start_nums]

        temp_rows_list = []  # New list to hold temporary DataFrames

        # Iterate over each number in start_nums and apply the filter
        for start_num in start_nums:
            # Get the rows in 'data_gebruik' where 'act1code' starts with 'start_num'
            temp_rows = self.data_gebruik[self.data_gebruik['act1code'].astype(str).str.startswith(str(start_num))]

            if not temp_rows.empty:
                temp_rows_list.append(temp_rows)  # Add the DataFrame to the list
            else:
                print(f"No act1code starting with {start_num} was found in the data_gebruik dataframe.")

        # Concatenate all temporary DataFrames in the list
        matched_rows_gebruik = pd.concat(temp_rows_list) if temp_rows_list else pd.DataFrame()

        if not matched_rows_gebruik.empty:
            # Get the unique 'aanzien_id' from the matched rows
            valid_aanzien_id = matched_rows_gebruik['aanzien_id'].unique()

            # Filter 'data_aanzien' for rows whose 'bronsleutel' is in 'valid_aanzien_id'
            original_rows = len(self.data_aanzien)
            self.data_aanzien = self.data_aanzien[self.data_aanzien['bronsleutel'].isin(valid_aanzien_id)]
            filtered_rows = len(self.data_aanzien)
            removed_rows = original_rows - filtered_rows
            
            # Fix: use dictionary format for consistency
            self.history.append({"action": f"filter SBI starting with {start_nums}", 
                                "rows_removed": removed_rows, 
                                "rows_remaining": filtered_rows})

            print(f"Filtering on act1code starting with {start_nums}")
            print(f"Removed {removed_rows} rows, {filtered_rows} rows remaining.")

    def prepare_dataframe(self, add_A=False):
        # Copy the dataframe to avoid modifying the original data
        df = self.HUP.copy()

        if add_A:
            df = pd.concat([df, self.original_data[self.original_data['risico_classificatie'] == 'A']])

        # Find the column name where value is 1 among the columns ending with 'functie'
        func_cols = [col for col in df.columns if col.endswith('functie')]
        df[func_cols] = df[func_cols].apply(pd.to_numeric, errors='coerce')

        # Get the column name which has maximum value in each row among the columns ending with 'functie'
        df['functie'] = df[func_cols].idxmax(axis=1)
        df['functie'] = df['functie'].apply(lambda x: x if pd.notna(x) and df.loc[df['functie'] == x, x].all() else '')

        # Combine straatnaam, huisnr and huistoevg columns into a single column
        df['adres'] = df['straatnaam'] + ' ' + df['huisnr'].apply(lambda x: str(int(x)) if pd.notna(x) else '') + df[
            'huisletter'].apply(
            lambda x: f'-{x}' if pd.notna(x) else '') + df['huistoevg'].apply(lambda x: f'{x}' if pd.notna(x) else '')

        # Filter self.data_gebruik for non-empty 'naam_vol' and take the first occurrence for each 'aanzien_id'
        data_gebruik_filtered = self.data_gebruik.loc[self.data_gebruik['naam_vol'].notna()].drop_duplicates(
            'aanzien_id', keep='first')

        # Map 'bronsleutel' from df to 'naam_vol' from data_gebruik_filtered
        df['Bouwwerk'] = df['bronsleutel'].map(data_gebruik_filtered.set_index('aanzien_id')['naam_vol'])

        # Filter self.data_gebruik for non-empty 'personen' and take the first occurrence for each 'aanzien_id'
        data_gebruik_filtered_personen = self.data_gebruik.loc[self.data_gebruik['personen'].notna()].drop_duplicates(
            'aanzien_id', keep='first')

        # Map 'bronsleutel' from df to 'personen' from data_gebruik_filtered_personen
        df['personen'] = df['bronsleutel'].map(data_gebruik_filtered_personen.set_index('aanzien_id')['personen'])

        # Filter self.data_gebruik for non-empty 'act1code' and take the first occurrence for each 'aanzien_id'
        data_gebruik_filtered_act1code = self.data_gebruik.loc[self.data_gebruik['act1code'].notna()].drop_duplicates(
            'aanzien_id', keep='first')

        # Map 'bronsleutel' from df to 'act1code' and 'act1omschr' from data_gebruik_filtered_act1code
        df['SBI1'] = df['bronsleutel'].map(data_gebruik_filtered_act1code.set_index('aanzien_id')['act1code'])
        df['act1omschr'] = df['bronsleutel'].map(data_gebruik_filtered_act1code.set_index('aanzien_id')['act1omschr'])

        # Check for duplicates.
        df.drop_duplicates(subset=['id'], keep='first', inplace=True)

        # Create new dataframe with selected and new columns
        new_df = pd.DataFrame()
        # New column order as specified
        new_df['Naam Bouwwerk'] = df['Bouwwerk']  # Renamed from 'Bouwwerk'
        new_df['Functie'] = df['functie']
        new_df['Risico'] = df['risico_classificatie']
        new_df['Wijziging Naam'] = None  # Previously 'Moet zijn'
        new_df['Adres'] = df['adres']
        new_df['Postcode'] = df['pc6']
        new_df['Gemeente'] = df['gemnaam']
        new_df['Checklist verstuurd?'] = None  # New column
        new_df['Checklist uitgevoerd?'] = None  # New column
        new_df['Datum uitgevoerd'] = None  # New column
        new_df['Personen'] = df['personen']
        new_df['SBI1'] = df['SBI1']
        new_df['SBI Omschrijving'] = df['act1omschr']
        new_df['Bouwjaar'] = df['bouwjaar']
        new_df['Bouwlagen'] = df['bouwlagen']
        new_df['Pandhoogte'] = df['pandhoogte']
        new_df['x'] = df['x']
        new_df['y'] = df['y']
        # New columns can be added here
        return new_df

    def insert_dataframe_into_excel(self, template_path, sheet_name, start_row, output_path=None, add_A=False, remove_no_name=False):
        """
        Insert dataframe into excel template at specified location.
        
        Args:
            template_path: Path to the Excel template
            sheet_name: Name of the sheet to insert data into
            start_row: Row number to start inserting data
            output_path: Optional path for the output file, if None a path will be generated
            add_A: Whether to include risk class A items
            remove_no_name: Whether to remove items without a name
            
        Returns:
            Path to the saved Excel file
        """
        df = self.prepare_dataframe(add_A)

        if remove_no_name:
            df = df.dropna(subset=['Bouwwerk'])

        # Load the workbook and select the specified worksheet
        workbook = load_workbook(filename=template_path)
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"{sheet_name} not in workbook. Available sheets are: {workbook.sheetnames}")

        sheet = workbook[sheet_name]

        # Insert blank rows to make space for DataFrame
        num_rows_df = len(df)
        sheet.insert_rows(start_row, num_rows_df)

        # Convert the dataframe to a list of rows and insert into Excel
        rows = dataframe_to_rows(df, index=False, header=False)
        for r_idx, row in enumerate(rows, start_row):
            # Iterate over all cells in the row
            for c_idx, value in enumerate(row, 1):
                # Insert the value into the cell
                sheet.cell(row=r_idx, column=c_idx, value=value)

        # Use the provided output_path if given, otherwise create one
        if output_path is None:
            # Prepare the new file name with date and time
            datetime_string = datetime.now().strftime("%d-%m-%Y_%H-%M")
            output_path = get_executable_relative_path("HUP", f"HUP-{datetime_string}.xlsx")
        
        # Make sure the directory exists
        os.makedirs(os.path.dirname(output_path), exist_ok=True)
        
        # Save the workbook
        workbook.save(output_path)
        print(f"Saved to {output_path}")
        return output_path
